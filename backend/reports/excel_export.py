from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

HEADER_FILL = PatternFill(
    start_color="1F3B57",
    end_color="1F3B57",
    fill_type="solid"
)

HEADER_FONT = Font(color="FFFFFF", bold=True)

PROPERTY_TYPE_LABELS = {
    'apartment': 'آپارتمان',
    'villa': 'ویلا',
    'land': 'زمین',
    'office': 'اداری',
    'shop': 'مغازه',
}

REQUEST_TYPE_LABELS = {
    'buy': 'خرید',
    'rent': 'اجاره',
    'sell': 'فروش',
    'mortgage': 'رهن',
}

FORMULA_PREFIXES = ('=', '+', '-', '@', '\t', '\r')


def _sanitize_cell(value):

    if isinstance(value, str) and value.startswith(FORMULA_PREFIXES):
        return "'" + value

    return value


def _sanitize_row(row):

    return [_sanitize_cell(value) for value in row]


def _write_header(ws, headers):

    ws.append(_sanitize_row(headers))

    for cell in ws[1]:

        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")


def _autofit(ws):

    for column_cells in ws.columns:

        length = max(
            len(str(cell.value)) if cell.value is not None else 0
            for cell in column_cells
        )

        ws.column_dimensions[column_cells[0].column_letter].width = max(12, length + 4)


def build_report_excel(report_data, start_date, end_date):

    wb = Workbook()

    ws_sales = wb.active
    ws_sales.title = "فروش و اجاره ماهانه"
    ws_sales.sheet_view.rightToLeft = True

    _write_header(ws_sales, [
        "ماه", "تعداد فروش", "مبلغ فروش (تومان)", "تعداد اجاره", "مبلغ اجاره (تومان)"
    ])

    for row in report_data['sales']:

        ws_sales.append(_sanitize_row([
            row['month'],
            row['sale_count'],
            row['sale_amount'],
            row['rent_count'],
            row['rent_amount'],
        ]))

    _autofit(ws_sales)

    ws_agents = wb.create_sheet("عملکرد مشاوران")
    ws_agents.sheet_view.rightToLeft = True

    _write_header(ws_agents, [
        "نام مشاور", "تعداد بازدید", "تعداد قرارداد", "مبلغ کل قراردادها (تومان)"
    ])

    for row in report_data['agents']:

        ws_agents.append(_sanitize_row([
            row['agent_name'],
            row['visits_count'],
            row['contracts_count'],
            row['contracts_amount'],
        ]))

    _autofit(ws_agents)

    ws_customers = wb.create_sheet("آمار مشتریان")
    ws_customers.sheet_view.rightToLeft = True

    customers = report_data['customers']

    _write_header(ws_customers, ["شاخص", "مقدار"])

    ws_customers.append(_sanitize_row(["تعداد کل مشتریان", customers['total']]))
    ws_customers.append(_sanitize_row(["تبدیل‌شده به مشتری قطعی", customers['converted']]))
    ws_customers.append(_sanitize_row(["نرخ تبدیل (٪)", customers['conversion_rate']]))
    ws_customers.append([])
    ws_customers.append(_sanitize_row(["نوع درخواست", "تعداد"]))

    for row in customers['by_request_type']:

        label = REQUEST_TYPE_LABELS.get(row['request_type'], row['request_type'])

        ws_customers.append(_sanitize_row([label, row['count']]))

    _autofit(ws_customers)

    ws_prices = wb.create_sheet("میانگین قیمت ملک")
    ws_prices.sheet_view.rightToLeft = True

    _write_header(ws_prices, [
        "نوع ملک", "تعداد", "میانگین قیمت (تومان)", "میانگین متراژ"
    ])

    for row in report_data['property_prices']:

        label = PROPERTY_TYPE_LABELS.get(row['property_type'], row['property_type'])

        ws_prices.append(_sanitize_row([
            label,
            row['count'],
            row['avg_price'],
            row['avg_area'],
        ]))

    _autofit(ws_prices)

    buffer = BytesIO()

    wb.save(buffer)

    buffer.seek(0)

    return buffer